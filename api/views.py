from django.shortcuts import render
from django.http import JsonResponse
from rest_framework import viewsets, status
from .models import appsecObservation, vaptPlugin, vaptObservation, scrObservation
from .serializers import appsecSerializer, vaptPluginSerializer, vaptSerializer, scrSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from openpyxl import load_workbook
import os
from reportMakerAPI.settings import BASE_DIR

class appsecViewSet(viewsets.ModelViewSet):
    queryset = appsecObservation.objects.all()
    serializer_class = appsecSerializer

class vaViewSet(viewsets.ModelViewSet):
    queryset = vaptObservation.objects.all()
    serializer_class = vaptSerializer

class scrViewSet(viewsets.ModelViewSet):
    queryset = scrObservation.objects.all()
    serializer_class = scrSerializer

@api_view(['POST'])
def sendVAPTObs(request):
    if request.method == "POST":
        requested_plugins = request.data["plugins"]
        requested_observations = vaptPlugin.objects.filter(pluginID__in=requested_plugins)
        serialied_obs = vaptPluginSerializer(requested_observations, many = True)
        return Response(serialied_obs.data)
    return Response({"error": "Incorrect method"})

def fillObservationsAppsec(request):
    if request.user.is_authenticated and request.user.is_superuser:
        wb = load_workbook(os.path.join(BASE_DIR, 'obs.xlsx'))
        ws_appsec = wb['appsec']

        # APPSEC
        for row in ws_appsec.iter_rows(min_row = 2, min_col = 1, max_col = 6):
            appsecObservation.objects.create(observation=row[0].value, detOb=row[1].value, criticality=row[2].value, risk=row[3].value, recommendation=row[4].value, abbr=row[5].value).save()

        wb.close()
        return JsonResponse({'result': 'Database Filled'})
    return JsonResponse({'result': 'Not Allowed'})

def fillObservationsVA(request):
    if request.user.is_authenticated and request.user.is_superuser:
        wb = load_workbook(os.path.join(BASE_DIR, 'obs.xlsx'))
        ws_va = wb['va']

        # VULNERABILITY ASSESSMENT
        for row in ws_va.iter_rows(min_row = 2, min_col = 1, max_col = 6):
            vaptObservation.objects.create(observation=row[0].value, detOb=row[1].value, criticality=row[2].value, risk=row[3].value, recommendation=row[4].value, abbr=row[5].value).save()

        wb.close()

        return JsonResponse({'result': 'Database Filled'})
    return JsonResponse({'result': 'Not Allowed'})

def fillObservationsSCR(request):
    if request.user.is_authenticated and request.user.is_superuser:
        wb = load_workbook(os.path.join(BASE_DIR, 'obs.xlsx'))
        ws_scr = wb['scr']

        # SOURCE CODE REVIEW
        for row in ws_scr.iter_rows(min_row = 2, min_col = 1, max_col = 7):
            scrObservation.objects.create(observation=row[0].value, detOb=row[1].value, criticality=row[2].value, risk=row[3].value, recommendation=row[4].value, abbr=row[5].value, language=row[6].value).save()
        wb.close()
        return JsonResponse({'result': 'Database Filled'})
    return JsonResponse({'result': 'Not Allowed'})
